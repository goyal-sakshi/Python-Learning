from openpyxl import load_workbook
import itertools

file_path = '/home/sa.goyal/Desktop/excel.xlsx'
file_obj = load_workbook(file_path)

dict_color = {
                'FFFF9900' : 'fail',      # orange
                'FF993300' : 'mid',       # red
                'FFFFFF00' : 'pass',      # yellow
                'FF008080' : 'fine'       # green
            }

obj = []
sheet = []
for i in range(0,2):
    obj.append(file_obj.get_sheet_names()[i])
    sheet.append(file_obj.get_sheet_by_name(obj[i]))

l1, l2, s1, s2 = [], [], [], []

for line in sheet[0]['D2':'D5']:
    for column in line:
        l1.append(column.fill.bgColor.rgb)


for line in sheet[1]['G2':'G5']:
    for column in line:
        l2.append(column.fill.bgColor.rgb)
        
for i in l1:
    if i in dict_color:
        s1.append(dict_color[i])
    else:
        s1.append("No color")

for i in l2:
    if i in dict_color:
        s2.append(dict_color[i])
    else:
        s2.append("No color")

sheet[0].cell(row=1, column=5).value = 'color_code'
sheet[1].cell(row=1, column=8).value = 'color_code'
r,j = 2, 0
for a,b in itertools.izip(s1,s2):
    sheet[0].cell(row=r, column=5).value = s1[j]
    sheet[1].cell(row=r, column=8).value = s2[j]
    r += 1
    j += 1
    
file_obj.save(file_path)
